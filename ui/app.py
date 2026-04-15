from __future__ import annotations

import html
import os
import sys
import tempfile
from collections import OrderedDict
from datetime import datetime

_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)
os.chdir(_PROJECT_ROOT)

import gradio as gr
import pandas as pd

from engine.config_loader import config_loader
from engine.licensor_context import build_context
from engine.reporter import build_output
from engine.validator import run_validation

_COL_BRAND_COLLECTION = "Brand/Collection"
_COL_LICENSING_STATUS = "Licensing Status"
_PHRASE_DIRECT_TO_RETAIL = "direct to retail"
_PHRASE_CONCEPT_READY = "concept ready for license submission"

# Excel preview in UI (first sheet only; row cap for large files)
_PREVIEW_ROW_CAP = 300


def _uploaded_file_path(file_obj) -> str | None:
    if file_obj is None:
        return None
    if isinstance(file_obj, str):
        return file_obj if os.path.isfile(file_obj) else None
    if isinstance(file_obj, dict):
        path = file_obj.get("path") or file_obj.get("name")
        path = str(path) if path else None
        return path if path and os.path.isfile(path) else None
    path = getattr(file_obj, "name", None)
    return path if path and os.path.isfile(path) else None


def on_upload_file_change(file_obj):
    """Show preview CTA when a file is present; hide popup when upload cleared."""
    if _uploaded_file_path(file_obj) is None:
        return gr.update(visible=False), gr.update(visible=False)
    return gr.update(visible=True), gr.update(visible=False)


def open_upload_preview(file_obj):
    if file_obj is None:
        raise gr.Error("Upload a Centric export (.xlsx) first.")
    path = _uploaded_file_path(file_obj)
    if not path:
        raise gr.Error("Could not read the uploaded file path.")
    try:
        df = pd.read_excel(path, nrows=_PREVIEW_ROW_CAP)
    except Exception as e:
        raise gr.Error(f"Could not preview this spreadsheet: {e}") from e
    name = os.path.basename(path)
    safe_name = html.escape(name)
    # Use HTML (not Markdown backticks): inline <code> from the theme was unreadable.
    caption_html = (
        f'<div class="upload-preview-caption" style="margin:0 0 12px 0;font-size:14px;'
        f'line-height:1.55;color:#0f172a;">'
        f"<strong>File:</strong> {safe_name} · <strong>{df.shape[1]}</strong> columns · "
        f"showing the <strong>first {len(df)}</strong> row(s) from the first sheet "
        f"(preview capped at {_PREVIEW_ROW_CAP} rows).</div>"
    )
    return gr.update(visible=True), df, caption_html


def close_upload_preview_modal():
    return gr.update(visible=False)


def apply_default_centric_filters(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Centric export defaults: drop Direct to Retail brands; keep only
    Concept Ready For License Submission. Preserves DataFrame index for
    validator/reporter row alignment.
    """
    original_rows = len(df)
    if _COL_BRAND_COLLECTION not in df.columns or _COL_LICENSING_STATUS not in df.columns:
        return df, {
            "applied": False,
            "original_rows": original_rows,
            "filtered_rows": original_rows,
            "dropped_dtr": 0,
            "dropped_status": 0,
        }

    s_brand = df[_COL_BRAND_COLLECTION].astype(str)
    s_status = df[_COL_LICENSING_STATUS].astype(str)
    is_direct_to_retail = s_brand.str.contains(_PHRASE_DIRECT_TO_RETAIL, case=False, na=False)
    is_concept_ready = s_status.str.contains(_PHRASE_CONCEPT_READY, case=False, na=False)
    df_filtered = df[~is_direct_to_retail & is_concept_ready]

    dropped_dtr = int(is_direct_to_retail.sum())
    dropped_status = int((~is_concept_ready).sum())

    return df_filtered, {
        "applied": True,
        "original_rows": original_rows,
        "filtered_rows": len(df_filtered),
        "dropped_dtr": dropped_dtr,
        "dropped_status": dropped_status,
    }


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------
config_loader.load_all()

_licensor_options = config_loader.available_licensors()
_licensor_map = {display: lid for lid, display in _licensor_options}
_licensor_names = [d for _, d in _licensor_options]


# ---------------------------------------------------------------------------
# Helper — group rule results by their parent rule group
# ---------------------------------------------------------------------------
def _group_results(row_results: list[dict], rules_map: dict) -> OrderedDict:
    """
    Returns OrderedDict of group_name -> list of sub-rule result dicts.
    Preserves insertion order so groups appear in evaluation order.
    """
    groups: OrderedDict = OrderedDict()
    for r in row_results:
        rule_def = rules_map.get(r["rule_id"])
        gname = rule_def.group_name if rule_def else "Other"
        gid = rule_def.group if rule_def else ""
        key = (gid, gname)
        if key not in groups:
            groups[key] = []
        groups[key].append(r)
    return groups


# ---------------------------------------------------------------------------
# Validation logic
# ---------------------------------------------------------------------------
def run(licensor_display: str, file_obj):
    if file_obj is None:
        raise gr.Error("Upload a Centric export (.xlsx) first.")

    licensor_id = _licensor_map.get(licensor_display)
    if not licensor_id:
        raise gr.Error("Select a valid licensor.")

    df = pd.read_excel(file_obj.name if hasattr(file_obj, "name") else file_obj)
    df, filter_meta = apply_default_centric_filters(df)
    if filter_meta["applied"] and filter_meta["filtered_rows"] == 0:
        raise gr.Error(
            "No rows left after default Centric filters "
            "(exclude Brand/Collection containing Direct to Retail; "
            "Licensing Status must include Concept Ready For License Submission)."
        )

    context = build_context(licensor_id)
    v_results = run_validation(df, context)
    _pass_df, _error_df, summary = build_output(df, v_results)

    rules_map = context["rules"]

    total = summary["total"]
    pass_count = summary["pass_count"]
    fail_count = summary["fail_count"]
    pass_pct = summary["pass_pct"]

    # -- Build rule stats ---------------------------------------------------
    from collections import Counter as _Counter

    group_stats: OrderedDict = OrderedDict()
    rule_stats: OrderedDict = OrderedDict()
    for vr in v_results:
        for r in vr["results"]:
            rid = r["rule_id"]
            st = r["status"]
            if rid not in rule_stats:
                rule_stats[rid] = _Counter()
            rule_stats[rid][st] += 1
            rd = rules_map.get(rid)
            gid = rd.group if rd else rid.split(".")[0]
            if gid not in group_stats:
                group_stats[gid] = _Counter()
            group_stats[gid][st] += 1

    # -- Build style-centric HTML ------------------------------------------
    html_parts = []

    if filter_meta["applied"]:
        orig = filter_meta["original_rows"]
        filt = filter_meta["filtered_rows"]
        html_parts.append(f"""
    <div class="centric-filter-banner" role="status" style="background:#ffffff;border:1px solid #94a3b8;border-left:5px solid #0066cc;
         border-radius:10px;padding:16px 20px;margin-bottom:18px;line-height:1.55;
         box-shadow:0 2px 6px rgba(15,23,42,0.08);color:#000000;">
      <div style="font-weight:800;font-size:15px;color:#000000;margin-bottom:10px;letter-spacing:-0.01em;">
        Default Centric filters applied
      </div>
      <ul style="margin:0;padding-left:22px;color:#000000;font-size:14px;">
        <li style="margin-bottom:6px;">Excluded rows where <strong style="color:#000000;">Brand/Collection</strong> contains
            <strong style="color:#000000;">Direct to Retail</strong>.</li>
        <li>Included only rows where <strong style="color:#000000;">Licensing Status</strong> contains
            <strong style="color:#000000;">Concept Ready For License Submission</strong>.</li>
      </ul>
      <div style="margin-top:12px;padding-top:12px;border-top:1px solid #e2e8f0;font-size:14px;color:#000000;font-weight:600;">
        Showing <span style="color:#000000;">{filt}</span> of <span style="color:#000000;">{orig}</span> rows from the upload.
      </div>
    </div>
    """)

    html_parts.append(f"""
    <div style="display:flex; gap:16px; margin-bottom:18px; flex-wrap:wrap;">
      <div style="flex:1; min-width:120px; background:#f8f9fa; border-radius:12px;
           padding:16px; text-align:center; border:1px solid #e9ecef;">
        <div style="font-size:26px; font-weight:700; color:#333;">{total}</div>
        <div style="font-size:12px; color:#666; margin-top:2px;">Total Styles</div>
      </div>
      <div style="flex:1; min-width:120px; background:#f0faf4; border-radius:12px;
           padding:16px; text-align:center; border:1px solid #c3e6cb;">
        <div style="font-size:26px; font-weight:700; color:#28a745;">{pass_count}</div>
        <div style="font-size:12px; color:#666; margin-top:2px;">Passed</div>
      </div>
      <div style="flex:1; min-width:120px; background:#fff5f5; border-radius:12px;
           padding:16px; text-align:center; border:1px solid #f5c6cb;">
        <div style="font-size:26px; font-weight:700; color:#dc3545;">{fail_count}</div>
        <div style="font-size:12px; color:#666; margin-top:2px;">Failed</div>
      </div>
      <div style="flex:1; min-width:120px; background:#f8f9fa; border-radius:12px;
           padding:16px; text-align:center; border:1px solid #e9ecef;">
        <div style="font-size:26px; font-weight:700; color:#0066cc;">{pass_pct}%</div>
        <div style="font-size:12px; color:#666; margin-top:2px;">Pass Rate</div>
      </div>
    </div>
    """)

    # -- Rule stats table --------------------------------------------------
    stats_rows_html = ""
    prev_gid = ""
    for rid, counts in rule_stats.items():
        rd = rules_map.get(rid)
        gid = rd.group if rd else rid.split(".")[0]
        gname = rd.group_name if rd else gid

        if gid != prev_gid:
            g_counts = group_stats.get(gid, _Counter())
            gp = g_counts.get("PASS", 0)
            gf = g_counts.get("FAIL", 0)
            gs = g_counts.get("SKIP", 0)
            gt = gp + gf + gs
            gp_pct = round(gp / gt * 100) if gt else 0
            bar_color = "#28a745" if gp_pct > 70 else "#e67e22" if gp_pct > 30 else "#dc3545"
            stats_rows_html += f"""
            <tr style="background:#edf2f7;border-top:2px solid #cbd5e0;">
              <td style="padding:6px 10px;font-weight:700;font-size:12px;color:#1a202c;">{gid}</td>
              <td style="padding:6px 10px;font-size:12px;font-weight:600;color:#1a202c;">{gname}</td>
              <td style="padding:6px 10px;text-align:center;font-size:12px;color:#22863a;font-weight:700;">{gp}</td>
              <td style="padding:6px 10px;text-align:center;font-size:12px;color:#cb2431;font-weight:700;">{gf}</td>
              <td style="padding:6px 10px;text-align:center;font-size:12px;color:#b08800;font-weight:700;">{gs}</td>
              <td style="padding:6px 8px;width:90px;">
                <div style="background:#e2e8f0;border-radius:3px;height:10px;overflow:hidden;">
                  <div style="background:{bar_color};height:100%;width:{gp_pct}%;border-radius:3px;"></div>
                </div>
              </td>
            </tr>"""
            prev_gid = gid

        rp = counts.get("PASS", 0)
        rf = counts.get("FAIL", 0)
        rs = counts.get("SKIP", 0)
        rname = rd.name if rd else rid
        stats_rows_html += f"""
        <tr style="background:#f7fafc;border-bottom:1px solid #e2e8f0;">
          <td style="padding:5px 10px 5px 26px;font-size:12px;color:#4a5568;">{rid}</td>
          <td style="padding:5px 10px;font-size:12px;color:#2d3748;">{rname}</td>
          <td style="padding:5px 10px;text-align:center;font-size:12px;color:#22863a;font-weight:600;">{rp}</td>
          <td style="padding:5px 10px;text-align:center;font-size:12px;color:#cb2431;font-weight:600;">{rf}</td>
          <td style="padding:5px 10px;text-align:center;font-size:12px;color:#b08800;font-weight:600;">{rs}</td>
          <td style="padding:5px 8px;"></td>
        </tr>"""

    html_parts.append(f"""
    <details style="margin-bottom:18px;" open>
      <summary style="cursor:pointer;font-size:14px;font-weight:700;color:#333;
               padding:8px 0;user-select:none;">
        Rule Statistics
      </summary>
      <table style="width:100%;border-collapse:collapse;font-size:12px;
             border:1px solid #e2e8f0;border-radius:8px;overflow:hidden;margin-top:6px;
             background:#fff;">
        <thead>
          <tr style="background:#edf2f7;">
            <th style="padding:8px 10px;text-align:left;font-size:12px;color:#2d3748;font-weight:700;width:55px;">Rule</th>
            <th style="padding:8px 10px;text-align:left;font-size:12px;color:#2d3748;font-weight:700;">Name</th>
            <th style="padding:8px 10px;text-align:center;font-size:12px;color:#22863a;font-weight:700;width:55px;">Pass</th>
            <th style="padding:8px 10px;text-align:center;font-size:12px;color:#cb2431;font-weight:700;width:55px;">Fail</th>
            <th style="padding:8px 10px;text-align:center;font-size:12px;color:#b08800;font-weight:700;width:55px;">Skip</th>
            <th style="padding:8px 8px;font-size:12px;color:#2d3748;font-weight:700;width:90px;">Pass %</th>
          </tr>
        </thead>
        <tbody>
          {stats_rows_html}
        </tbody>
      </table>
    </details>
    """)

    html_parts.append(f"""
    <div style="display:flex;gap:8px;margin-bottom:16px;align-items:center;">
      <span style="font-size:14px;font-weight:600;color:#444;margin-right:4px;">Filter:</span>
      <button onclick="document.querySelectorAll('.style-card').forEach(c=>c.style.display='');
              document.querySelectorAll('.filter-btn').forEach(b=>{{b.style.background='#fff';b.style.color='#555';}});
              this.style.background='#0066cc';this.style.color='#fff';"
              class="filter-btn"
              style="padding:6px 18px;border-radius:20px;border:1px solid #0066cc;
              background:#0066cc;color:#fff;cursor:pointer;font-size:13px;font-weight:600;">
        All ({total})
      </button>
      <button onclick="document.querySelectorAll('.style-card').forEach(c=>c.style.display=
              c.dataset.status==='PASS'?'':'none');
              document.querySelectorAll('.filter-btn').forEach(b=>{{b.style.background='#fff';b.style.color='#555';}});
              this.style.background='#28a745';this.style.color='#fff';"
              class="filter-btn"
              style="padding:6px 18px;border-radius:20px;border:1px solid #28a745;
              background:#fff;color:#555;cursor:pointer;font-size:13px;font-weight:600;">
        Passed ({pass_count})
      </button>
      <button onclick="document.querySelectorAll('.style-card').forEach(c=>c.style.display=
              c.dataset.status==='FAIL'?'':'none');
              document.querySelectorAll('.filter-btn').forEach(b=>{{b.style.background='#fff';b.style.color='#555';}});
              this.style.background='#dc3545';this.style.color='#fff';"
              class="filter-btn"
              style="padding:6px 18px;border-radius:20px;border:1px solid #dc3545;
              background:#fff;color:#555;cursor:pointer;font-size:13px;font-weight:600;">
        Failed ({fail_count})
      </button>
    </div>
    """)

    for vr in v_results:
        style_num = vr["Style #"]
        prop = vr["Property"]
        item_type = vr["Item Type"]
        customer = vr["Customer"]
        overall = vr["overall_status"]
        row_results = vr["results"]

        rules_passed = sum(1 for r in row_results if r["status"] == "PASS")
        rules_failed = sum(1 for r in row_results if r["status"] == "FAIL")
        rules_skipped = sum(1 for r in row_results if r["status"] == "SKIP")

        is_pass = overall == "PASS"
        card_bg = "#f0faf4" if is_pass else "#fff5f5"
        border_color = "#28a745" if is_pass else "#dc3545"
        badge_bg = "#28a745" if is_pass else "#dc3545"
        badge_text = "PASS" if is_pass else "FAIL"

        score_parts = []
        if rules_passed:
            score_parts.append(f'<span style="color:#28a745;font-weight:600;">{rules_passed} passed</span>')
        if rules_failed:
            score_parts.append(f'<span style="color:#dc3545;font-weight:600;">{rules_failed} failed</span>')
        if rules_skipped:
            score_parts.append(f'<span style="color:#856404;font-weight:600;">{rules_skipped} skipped</span>')
        score_html = " &middot; ".join(score_parts)

        card_html = f"""
        <div class="style-card" data-status="{overall}"
             style="background:{card_bg};border-left:5px solid {border_color};
             border-radius:10px;padding:18px 22px;margin-bottom:12px;
             box-shadow:0 1px 3px rgba(0,0,0,0.06);">
          <div style="display:flex;justify-content:space-between;align-items:center;
               margin-bottom:10px;flex-wrap:wrap;gap:8px;">
            <div>
              <span style="font-size:17px;font-weight:700;color:#222;">{style_num}</span>
              <span style="color:#555;margin-left:14px;font-size:14px;">{prop}</span>
              <span style="color:#999;margin-left:10px;font-size:13px;">{item_type} &middot; {customer}</span>
            </div>
            <div style="display:flex;align-items:center;gap:12px;">
              <span style="font-size:12px;color:#555;">{score_html}</span>
              <span style="background:{badge_bg};color:white;padding:5px 16px;
                    border-radius:20px;font-size:13px;font-weight:700;
                    letter-spacing:0.5px;">{badge_text}</span>
            </div>
          </div>
          <div style="display:flex;flex-direction:column;gap:6px;">
        """

        grouped = _group_results(row_results, rules_map)
        for (gid, gname), sub_results in grouped.items():
            card_html += f"""
            <div style="font-size:12px;font-weight:700;color:#555;margin-top:6px;
                 text-transform:uppercase;letter-spacing:0.5px;">
              {gid}: {gname}
            </div>
            """
            for r in sub_results:
                status = r["status"]
                if status == "PASS":
                    icon, bg, color = "&#10003;", "#d4edda", "#155724"
                elif status == "FAIL":
                    icon, bg, color = "&#10007;", "#f8d7da", "#721c24"
                else:
                    icon, bg, color = "&#8674;", "#fff3cd", "#856404"

                reason_html = ""
                if r["reason"]:
                    reason_html = (
                        f'<span style="color:{color};font-size:12px;margin-left:10px;'
                        f'opacity:0.85;">&rarr; {r["reason"]}</span>'
                    )

                card_html += f"""
                <div style="background:{bg};border-radius:6px;padding:8px 14px;
                     display:flex;align-items:center;flex-wrap:wrap;gap:4px;margin-left:12px;">
                  <span style="font-weight:700;color:{color};min-width:55px;font-size:13px;">
                    {icon} {r["rule_id"]}
                  </span>
                  <span style="color:{color};font-size:13px;">{r["rule_name"]}</span>
                  {reason_html}
                </div>
                """

        card_html += "</div></div>"
        html_parts.append(card_html)

    results_html = "\n".join(html_parts)

    # -- Build single downloadable report ----------------------------------
    report_rows = []
    for vr in v_results:
        check_lines = []
        for r in vr["results"]:
            line = f"[{r['status']}] {r['rule_id']} - {r['rule_name']}"
            if r["reason"]:
                line += f" : {r['reason']}"
            check_lines.append(line)

        overall = vr["overall_status"]
        submission_msg = (
            "Ready for submission" if overall == "PASS" else "Not ready for submission"
        )
        report_rows.append({
            "Style #": vr["Style #"],
            "Property": vr["Property"],
            "Item Type": vr["Item Type"],
            "Customer": vr["Customer"],
            "Licensor": vr["Licensor"],
            "Status": overall,
            "Submission status": submission_msg,
            "Checks Passed": sum(1 for r in vr["results"] if r["status"] == "PASS"),
            "Checks Failed": sum(1 for r in vr["results"] if r["status"] == "FAIL"),
            "Checks Skipped": sum(1 for r in vr["results"] if r["status"] == "SKIP"),
            "Check Details": "\n".join(check_lines),
        })

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_path = os.path.join(tempfile.gettempdir(), f"validation_report_{ts}.xlsx")
    report_df = pd.DataFrame(report_rows)
    report_df.to_excel(report_path, index=False, engine="openpyxl")

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(report_path)
    ws = wb.active
    detail_col_letter = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Check Details":
            detail_col_letter = cell.column_letter
            break
    if detail_col_letter:
        ws.column_dimensions[detail_col_letter].width = 80
        for row in ws.iter_rows(min_row=2, min_col=ws[detail_col_letter + "1"].column,
                                max_col=ws[detail_col_letter + "1"].column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(report_path)

    return results_html, gr.update(value=report_path, visible=True)


# ---------------------------------------------------------------------------
# Config view
# ---------------------------------------------------------------------------
def build_config_html():
    parts = []
    for lid, display in _licensor_options:
        cfg = config_loader.get_config(lid)
        parts.append(f"""
        <div style="margin-bottom:24px;">
          <h3 style="margin:0 0 12px 0; color:#0066cc; font-size:20px;">{display}</h3>
          <table style="width:100%; border-collapse:collapse; font-size:14px;
                 border:1px solid #dee2e6; border-radius:8px; overflow:hidden;">
            <thead>
              <tr style="background:#e9ecef; text-align:left;">
                <th style="padding:12px 14px; border-bottom:2px solid #ced4da; color:#333; font-weight:700;">Group</th>
                <th style="padding:12px 14px; border-bottom:2px solid #ced4da; color:#333; font-weight:700;">ID</th>
                <th style="padding:12px 14px; border-bottom:2px solid #ced4da; color:#333; font-weight:700;">Rule Name</th>
                <th style="padding:12px 14px; border-bottom:2px solid #ced4da; color:#333; font-weight:700;">Enabled</th>
                <th style="padding:12px 14px; border-bottom:2px solid #ced4da; color:#333; font-weight:700;">Gate</th>
                <th style="padding:12px 14px; border-bottom:2px solid #ced4da; color:#333; font-weight:700;">Config File</th>
              </tr>
            </thead>
            <tbody>
        """)

        prev_group = ""
        for r in cfg.rules:
            group_cell = ""
            if r.group != prev_group:
                group_cell = f"<strong>{r.group}</strong>: {r.group_name}"
                prev_group = r.group

            en_badge = ('<span style="color:#28a745; font-weight:600;">Enabled</span>'
                        if r.enabled else
                        '<span style="color:#dc3545; font-weight:600;">Disabled</span>')
            gate_badge = ('<span style="background:#e67e22; color:#fff; padding:3px 10px;'
                          ' border-radius:4px; font-size:12px; font-weight:600;">Gate</span>'
                          if r.is_gate else "\u2014")
            cf = r.config_file or "\u2014"
            parts.append(f"""
              <tr style="border-bottom:1px solid #e9ecef; background:#fff;">
                <td style="padding:10px 14px; color:#333;">{group_cell}</td>
                <td style="padding:10px 14px; font-weight:600; color:#333;">{r.id}</td>
                <td style="padding:10px 14px; color:#333;">{r.name}</td>
                <td style="padding:10px 14px;">{en_badge}</td>
                <td style="padding:10px 14px;">{gate_badge}</td>
                <td style="padding:10px 14px; font-size:12px; color:#555;">{cf}</td>
              </tr>
            """)
        parts.append("</tbody></table></div>")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Gradio UI
# ---------------------------------------------------------------------------
css = """
.gradio-container { max-width: 1100px !important; }
#results-area, #results-area > div, .full-scroll, .full-scroll > div {
    max-height: none !important;
    overflow: visible !important;
    height: auto !important;
}
/* Gradio Soft theme tints list body text; force black for the Centric filter callout */
.centric-filter-banner,
.centric-filter-banner * {
    color: #000000 !important;
}
/*
 * Preview popup: do NOT set display: flex !important on #upload-preview-modal — it
 * overrides Gradio's display:none when visible=False, so the overlay would always show.
 */
#upload-preview-modal {
    position: fixed !important;
    inset: 0 !important;
    z-index: 10050 !important;
    width: 100vw !important;
    max-width: 100vw !important;
    min-height: 100vh !important;
    margin: 0 !important;
    padding: 24px !important;
    box-sizing: border-box !important;
    background: rgba(15, 23, 42, 0.6) !important;
    /* Single scroll inside the card — nested overflow:auto here caused jitter */
    overflow: hidden !important;
}
#upload-preview-modal .upload-preview-card {
    width: min(1120px, 96vw);
    max-height: min(86vh, 900px);
    margin: 4vh auto;
    overflow-x: auto;
    overflow-y: auto;
    overscroll-behavior: contain;
    background: #ffffff !important;
    color: #0f172a !important;
    border-radius: 12px;
    padding: 16px 20px 20px;
    box-shadow: 0 24px 48px rgba(0, 0, 0, 0.28);
    border: 1px solid #cbd5e1;
}
/* Theme often uses light gray body text inside blocks; force readable contrast */
#upload-preview-modal .upload-preview-card,
#upload-preview-modal .upload-preview-card p,
#upload-preview-modal .upload-preview-card li,
#upload-preview-modal .upload-preview-card h1,
#upload-preview-modal .upload-preview-card h2,
#upload-preview-modal .upload-preview-card h3,
#upload-preview-modal .upload-preview-card th,
#upload-preview-modal .upload-preview-card td,
#upload-preview-modal .upload-preview-card span,
#upload-preview-modal .upload-preview-card label,
#upload-preview-modal .upload-preview-card .label-wrap,
#upload-preview-modal .upload-preview-card .prose,
#upload-preview-modal .upload-preview-card .prose * {
    color: #0f172a !important;
}
#upload-preview-modal .upload-preview-card button {
    color: #0f172a !important;
    background: #e2e8f0 !important;
    border: 1px solid #94a3b8 !important;
}
#upload-preview-modal .upload-preview-card .table-wrap,
#upload-preview-modal .upload-preview-card table,
#upload-preview-modal .upload-preview-card [class*="table"] {
    color: #0f172a !important;
    background: #ffffff !important;
}
#upload-preview-modal .upload-preview-card [role="gridcell"],
#upload-preview-modal .upload-preview-card [role="columnheader"] {
    color: #0f172a !important;
    background-color: #ffffff !important;
}
#upload-preview-modal .upload-preview-card .upload-preview-cell-wrap,
#upload-preview-modal .upload-preview-card .upload-preview-cell-wrap * {
    color: #0f172a !important;
}
/* Dataframe: light theme vars. Do not zebra-stripe with :nth-child — virtualizer
   reuses <tr> nodes while scrolling, so DOM row index != data row → flicker. */
#upload-preview-modal #upload-preview-dataframe {
    --body-text-color: #0f172a !important;
    --color-text-primary: #0f172a !important;
    --color-text-secondary: #334155 !important;
    --background-fill-primary: #ffffff !important;
    --border-color-primary: #e2e8f0 !important;
    --table-even-background-fill: #ffffff !important;
    --table-odd-background-fill: #ffffff !important;
    background: #ffffff !important;
    color: #0f172a !important;
}
#upload-preview-modal #upload-preview-dataframe svelte-virtual-table-viewport,
#upload-preview-modal #upload-preview-dataframe table,
#upload-preview-modal #upload-preview-dataframe thead,
#upload-preview-modal #upload-preview-dataframe tbody,
#upload-preview-modal #upload-preview-dataframe tfoot,
#upload-preview-modal #upload-preview-dataframe th,
#upload-preview-modal #upload-preview-dataframe td {
    background-color: #ffffff !important;
    color: #0f172a !important;
}
#upload-preview-modal #upload-preview-dataframe th,
#upload-preview-modal #upload-preview-dataframe td {
    border-color: #e2e8f0 !important;
}
#upload-preview-modal #upload-preview-dataframe thead th {
    background-color: #edf2f7 !important;
    color: #0f172a !important;
    font-weight: 600 !important;
}
#upload-preview-modal #upload-preview-dataframe input,
#upload-preview-modal #upload-preview-dataframe textarea {
    background: #ffffff !important;
    color: #0f172a !important;
}
#upload-preview-modal .upload-preview-caption,
#upload-preview-modal .upload-preview-caption * {
    color: #0f172a !important;
}
"""

with gr.Blocks(
    title="BioWorld Licensing Rule Engine",
    theme=gr.themes.Soft(primary_hue="blue", neutral_hue="slate"),
    css=css,
) as app:

    gr.Markdown(
        "# BioWorld Licensing Rule Engine\n"
        "Upload a Centric PLM export, pick a licensor, and validate every style "
        "against the configured licensing rules."
    )

    with gr.Tabs():
        with gr.Tab("Validate"):
            with gr.Column(visible=False, elem_id="upload-preview-modal") as upload_preview_modal:
                with gr.Column(elem_classes=["upload-preview-card"]):
                    with gr.Row():
                        gr.Markdown("### Preview: uploaded spreadsheet")
                        close_preview_btn = gr.Button("Close", variant="secondary", size="sm")
                    preview_caption_html = gr.HTML(value="")
                    preview_dataframe = gr.Dataframe(
                        label="Sheet data (read-only)",
                        interactive=False,
                        wrap=True,
                        height=520,
                        elem_id="upload-preview-dataframe",
                        elem_classes=["upload-preview-cell-wrap"],
                    )

            with gr.Row():
                licensor_dd = gr.Dropdown(
                    choices=_licensor_names,
                    value=_licensor_names[0] if _licensor_names else None,
                    label="Licensor",
                    scale=1,
                )
                file_upload = gr.File(
                    label="Upload Centric Export (.xlsx)",
                    file_types=[".xlsx"],
                    scale=2,
                )

            with gr.Row():
                preview_upload_btn = gr.Button(
                    "Preview uploaded data",
                    visible=False,
                    variant="secondary",
                    size="sm",
                )

            with gr.Row():
                run_btn = gr.Button("Run Validation", variant="primary", size="lg", scale=3)
                download_file = gr.File(
                    label="Download Report",
                    visible=False,
                    interactive=False,
                    scale=2,
                )

            results_html = gr.HTML(
                elem_id="results-area",
                value="<div style='text-align:center; padding:40px; color:#999;'>"
                      "Upload a file and click <b>Run Validation</b> to see results.</div>",
                elem_classes=["full-scroll"],
            )

            file_upload.change(
                fn=on_upload_file_change,
                inputs=[file_upload],
                outputs=[preview_upload_btn, upload_preview_modal],
            )

            preview_upload_btn.click(
                fn=open_upload_preview,
                inputs=[file_upload],
                outputs=[upload_preview_modal, preview_dataframe, preview_caption_html],
            )

            close_preview_btn.click(
                fn=close_upload_preview_modal,
                outputs=[upload_preview_modal],
            )

            run_btn.click(
                fn=run,
                inputs=[licensor_dd, file_upload],
                outputs=[results_html, download_file],
            )

        with gr.Tab("Config"):
            gr.Markdown("### Active Configuration\nRead-only view of rules loaded from `licensor_registry.yml`.")
            gr.HTML(value=build_config_html())


if __name__ == "__main__":
    app.launch(server_name="0.0.0.0", server_port=7860, share=True)
